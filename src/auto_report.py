"""
src/auto_report.py
Phase 4 ペーパートレード - 月次集計自動更新
実行順序: 4番目（最後）
"""
import sys
import os
sys.stdout.reconfigure(encoding="utf-8")

from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import sys as _sys
_sys.path.insert(0, str(Path(__file__).parent))
from notifier import send_notify, send_error_notify

_SCRIPT   = "auto_report.py"
TRADE_LOG = "logs/paper_trade_log.xlsx"
SCHED_LOG = "logs/scheduler_log.txt"
PDCA_LOG  = "logs/pdca_log.txt"


def log(msg: str) -> None:
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] [report] {msg}"
    print(line)
    os.makedirs("logs", exist_ok=True)
    try:
        with open(SCHED_LOG, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except PermissionError:
        pass


def detect_pdca_impulses(df: pd.DataFrame) -> list[str]:
    """今日の取引・シグナル状況から『変えたくなった衝動』を自動検出する"""
    today = datetime.now().date()
    entries = []

    # ── 損切り検出 ──────────────────────────────────────
    if not df.empty:
        today_df  = df[df["決済日"].dt.date == today]
        stops     = today_df[today_df["決済理由"].str.contains("損切り", na=False)]
        n_stops   = len(stops)
        if n_stops >= 2:
            codes = ", ".join(
                f"[{r['戦略']}]{r['銘柄コード']}" for _, r in stops.iterrows()
            )
            entries.append(
                f"[{today}] 同日{n_stops}件損切り（{codes}）"
                f"→ 同一相場環境での順張り集中リスクを実感。"
                f"同日エントリー上限を1件に絞りたくなるがOOS前提を壊すため Phase4凍結中・保留"
            )
        elif n_stops == 1:
            row     = stops.iloc[0]
            pct_str = (
                f"{float(row['損益率'])*100:+.1f}%"
                if pd.notna(row.get("損益率")) else ""
            )
            entries.append(
                f"[{today}] [{row['戦略']}]{row['銘柄コード']} 損切り{pct_str}"
                f" → ATR×2のSL幅を広げたくなるがOOS検証済みパラメータのため Phase4凍結中・保留"
            )

    # ── シグナルゼロ連続チェック ────────────────────────
    scanner_csv = Path("logs/scanner_log.csv")
    if scanner_csv.exists():
        try:
            sl_df       = pd.read_csv(scanner_csv, parse_dates=["Date"])
            last_dates  = sl_df["Date"].drop_duplicates().nlargest(7)
            recent      = sl_df[sl_df["Date"].isin(last_dates)]

            # 戦略別ゼロシグナル連続確認（F は別扱い）
            for strat_chk, reason_hint in [
                ("A", "RSIレンジかADX閾値を緩めたくなるがOOS検証済みのため"),
                ("E", "52週高値更新銘柄が減少中。パラメータ変更したくなるが"),
                ("F", "EPS成長≥20%の通期決算開示がない状態。EPS閾値を下げたくなるが"),
            ]:
                st_df = recent[recent.get("Strategy", pd.Series(dtype=str)) == strat_chk] \
                        if "Strategy" in recent.columns else pd.DataFrame()
                if st_df.empty:
                    continue
                daily_sig = st_df.groupby("Date")["Signal"].apply(
                    lambda x: (x.astype(str).str.strip() == "True").any()
                ).sort_index()
                zero_streak = 0
                for has_sig in reversed(daily_sig.values.tolist()):
                    if not has_sig:
                        zero_streak += 1
                    else:
                        break
                if zero_streak >= 5:
                    entries.append(
                        f"[{today}] 戦略{strat_chk}シグナルゼロ{zero_streak}営業日連続"
                        f" → {reason_hint} Phase4凍結中・保留"
                    )

            # 全戦略合計のゼロ連続（従来チェック）
            daily_sig_all = recent.groupby("Date")["Signal"].apply(
                lambda x: (x.astype(str).str.strip() == "True").any()
            ).sort_index()
            zero_streak_all = 0
            for has_sig in reversed(daily_sig_all.values.tolist()):
                if not has_sig:
                    zero_streak_all += 1
                else:
                    break
            if zero_streak_all >= 3:
                entries.append(
                    f"[{today}] 全戦略シグナルゼロ{zero_streak_all}営業日連続"
                    f" → 市場環境の変化を疑うが Phase4凍結中・保留"
                )
        except Exception:
            pass

    return entries


def append_pdca_log(entries: list[str]) -> None:
    if not entries:
        return
    today_str = datetime.now().strftime("%Y-%m-%d")
    os.makedirs("logs", exist_ok=True)
    with open(PDCA_LOG, "a", encoding="utf-8") as f:
        f.write(f"\n## {today_str}（自動記録）\n")
        for e in entries:
            f.write(e + "\n")


def read_config(wb) -> dict:
    ws = wb["設定"]
    config = {}
    for row in ws.iter_rows(min_row=3, max_row=15, values_only=True):
        if row[0] and row[1] is not None:
            config[str(row[0])] = row[1]
    return config


def load_completed_trades(ws_trade) -> pd.DataFrame:
    """決済済み（保有中以外）の全取引をDataFrameで返す"""
    rows = []
    cols = ["日付", "銘柄コード", "エントリー価格", "損切り価格", "利確価格",
            "ATR", "RSI", "ADX", "決済日", "決済価格", "損益円", "損益率", "決済理由", "戦略"]

    for row in ws_trade.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        reason = str(row[12]).strip() if row[12] is not None else ""
        if reason == "保有中" or reason == "":
            continue
        strategy = str(row[13]).strip() if len(row) > 13 and row[13] is not None else "A"
        rows.append(tuple(row[:13]) + (strategy,))

    if not rows:
        return pd.DataFrame(columns=cols)

    df = pd.DataFrame(rows, columns=cols)
    df["決済日"]     = pd.to_datetime(df["決済日"], errors="coerce")
    df["損益円"]     = pd.to_numeric(df["損益円"],     errors="coerce")
    df["損益率"]     = pd.to_numeric(df["損益率"],     errors="coerce")
    df["エントリー価格"] = pd.to_numeric(df["エントリー価格"], errors="coerce")
    return df.dropna(subset=["決済日", "損益円"])


def build_monthly_summary(df: pd.DataFrame, capital: float) -> pd.DataFrame:
    df["年月"] = df["決済日"].dt.to_period("M")

    monthly = (
        df.groupby("年月")
        .agg(
            取引数      = ("損益円", "count"),
            勝率        = ("損益円", lambda x: (x > 0).mean()),
            月次損益円  = ("損益円", "sum"),
        )
        .reset_index()
    )

    monthly["月次損益率"]       = monthly["月次損益円"] / capital
    monthly["月次ストップ発動"] = (monthly["月次損益円"] < -capital * 0.10).astype(int)
    monthly["年月"]             = monthly["年月"].astype(str)
    return monthly


def update_summary_sheet(ws_summary, monthly: pd.DataFrame) -> None:
    # データ行をクリア（ヘッダー行1は残す）
    for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row):
        for cell in row:
            cell.value = None

    for i, r in enumerate(monthly.itertuples(index=False), start=2):
        ws_summary.cell(row=i, column=1, value=r.年月)
        ws_summary.cell(row=i, column=2, value=int(r.取引数))
        ws_summary.cell(row=i, column=3, value=round(float(r.勝率), 4))
        ws_summary.cell(row=i, column=4, value=round(float(r.月次損益円), 0))
        ws_summary.cell(row=i, column=5, value=round(float(r.月次損益率), 6))
        ws_summary.cell(row=i, column=6, value=int(r.月次ストップ発動))


def main() -> None:
    log("auto_report.py 開始")

    if not Path(TRADE_LOG).exists():
        log(f"INFO: {TRADE_LOG} が見つかりません - スキップ")
        return

    wb          = load_workbook(TRADE_LOG)
    ws_trade    = wb["取引記録"]
    ws_summary  = wb["月次集計"]
    config      = read_config(wb)
    capital     = float(config.get("運用資金", 1_000_000))

    df = load_completed_trades(ws_trade)

    # 保有中銘柄数
    holding_count = sum(
        1 for row in ws_trade.iter_rows(min_row=2, values_only=True)
        if row[0] is not None and str(row[12]).strip() == "保有中"
    )

    # 日次レポート（完了取引ゼロでも常に送信）
    total_pnl    = float(df["損益円"].sum()) if not df.empty else 0.0
    current      = capital + total_pnl
    target_pnl   = capital * 0.02
    target_asset = capital + target_pnl
    achievement  = (total_pnl / target_pnl * 100) if target_pnl != 0 else 0.0

    # 戦略別損益集計
    strat_lines = []
    if not df.empty and "戦略" in df.columns:
        for st in ["A", "C", "D", "E", "F"]:
            sub = df[df["戦略"] == st]
            if not sub.empty:
                st_pnl = sub["損益円"].sum()
                st_wr  = (sub["損益円"] > 0).mean() * 100
                strat_lines.append(f"  [{st}] {len(sub)}件 WR{st_wr:.0f}% {st_pnl:+,.0f}円")

    # PDCA自動検出
    pdca_entries = detect_pdca_impulses(df)
    append_pdca_log(pdca_entries)
    if pdca_entries:
        log(f"INFO: PDCA自動記録 {len(pdca_entries)}件 → {PDCA_LOG}")

    # 通知本文
    pdca_lines = [f"・{e.split('→')[0].strip()}" for e in pdca_entries]
    daily_body = (
        f"保有中：{holding_count}銘柄\n"
        f"初期資金：{capital:,.0f}円\n"
        f"累計確定損益：{total_pnl:+,.0f}円\n"
        f"現在資産：{current:,.0f}円\n"
        f"目標（年利8%・3ヶ月）：{target_asset:,.0f}円\n"
        f"達成率：{achievement:.0f}%"
        + ("\n戦略別:\n" + "\n".join(strat_lines) if strat_lines else "")
        + ("\n\n📝 PDCA:\n" + "\n".join(pdca_lines) if pdca_lines else "")
    )
    send_notify("【ST】日次レポート", daily_body)
    log(f"INFO: 日次レポート送信  累計損益 {total_pnl:+,.0f}円  達成率 {achievement:.0f}%")

    if df.empty:
        log("INFO: 集計対象の完了取引なし")
        wb.close()
        return

    monthly = build_monthly_summary(df, capital)
    update_summary_sheet(ws_summary, monthly)
    wb.save(TRADE_LOG)

    # サマリーログ
    total_trades = int(monthly["取引数"].sum())
    log(f"INFO: 集計完了  {len(monthly)} ヶ月 / 計{total_trades}取引 / "
        f"累計損益 {total_pnl:+,.0f}円 → {TRADE_LOG}")

    for _, row in monthly.iterrows():
        stop = "★ストップ発動" if row["月次ストップ発動"] else ""
        log(f"  {row['年月']}  {int(row['取引数'])}件  "
            f"勝率{row['勝率']:.1%}  "
            f"{row['月次損益円']:+,.0f}円 ({row['月次損益率']:+.2%})  {stop}")


if __name__ == "__main__":
    try:
        main()
    except PermissionError as e:
        log(f"ERROR: PermissionError: {e}")
        send_error_notify(_SCRIPT, "PermissionError", str(e))
        sys.exit(1)
    except FileNotFoundError as e:
        log(f"ERROR: FileNotFoundError: {e}")
        send_error_notify(_SCRIPT, "FileNotFoundError", str(e))
        sys.exit(1)
    except ValueError as e:
        log(f"ERROR: ValueError: {e}")
        send_error_notify(_SCRIPT, "ValueError", str(e))
        sys.exit(1)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        log(f"ERROR: {type(e).__name__}: {e}\n{tb}")
        send_error_notify(_SCRIPT, type(e).__name__, str(e))
        sys.exit(1)
    finally:
        os.makedirs("logs", exist_ok=True)
        try:
            with open(SCHED_LOG, "a", encoding="utf-8") as f:
                f.write("========================================\n")
        except PermissionError:
            pass
