"""
src/auto_entry.py
Phase 4/5 エントリー自動記録（+ Phase 5 実注文）

LIVE_TRADING=false（デフォルト）: ペーパートレードのみ
LIVE_TRADING=true              : ペーパー記録 + kabu API 実注文
"""
import sys
import os
sys.stdout.reconfigure(encoding="utf-8")

import pandas as pd
from datetime import date, datetime
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook
import sys as _sys
_sys.path.insert(0, str(Path(__file__).parent))
from notifier import send_notify, send_error_notify

load_dotenv()

_SCRIPT      = "auto_entry.py"
SCANNER_LOG  = "logs/scanner_log.csv"
TRADE_LOG    = "logs/paper_trade_log.xlsx"
SCHED_LOG    = "logs/scheduler_log.txt"
MONTHLY_STOP = "logs/monthly_stop.txt"

MAX_POSITIONS    = 5
MAX_PER_STOCK    = 200_000
LIVE_TRADING     = os.getenv("LIVE_TRADING", "false").lower() == "true"
PHASE5_MAX_CAP   = float(os.getenv("PHASE5_MAX_CAPITAL", "500000"))


def log(msg: str) -> None:
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] [entry] {msg}"
    print(line)
    os.makedirs("logs", exist_ok=True)
    try:
        with open(SCHED_LOG, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except PermissionError:
        pass


def read_config(wb) -> dict:
    ws = wb["設定"]
    config = {}
    for row in ws.iter_rows(min_row=3, max_row=15, values_only=True):
        if row[0] and row[1] is not None:
            config[str(row[0])] = row[1]
    return config


def get_current_holdings(ws_trade) -> list[dict]:
    holdings = []
    for row in ws_trade.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        if str(row[12].value).strip() == "保有中":
            holdings.append({"code": str(row[1].value).strip()})
    return holdings


def get_next_empty_row(ws_trade) -> int:
    for row in ws_trade.iter_rows(min_row=2, max_row=500):
        if row[0].value is None:
            return row[0].row
    return ws_trade.max_row + 1


def load_today_signals() -> pd.DataFrame:
    if not Path(SCANNER_LOG).exists():
        log(f"WARN: {SCANNER_LOG} が見つかりません")
        return pd.DataFrame()

    df = pd.read_csv(SCANNER_LOG, dtype={"Code": str})
    today_str = date.today().isoformat()
    df = df[df["Date"] == today_str]
    df = df.drop_duplicates(subset=["Code"], keep="last")
    df = df[df["Signal"].astype(str).str.strip() == "True"]
    return df


def _place_live_buy(code4_or_5: str, shares: int, ref_price: float) -> str | None:
    """
    kabu API で現物成行買い注文を発行する。
    失敗してもペーパートレード記録は継続するため例外は吸収する。
    Returns: OrderId（成功時）または None（失敗時）
    """
    try:
        from broker_client import KabuClient, KabuClientError
        client = KabuClient()
        client.authenticate()

        # Phase 5 安全上限チェック
        positions = client.get_positions()
        current_value = sum(
            int(p.get("LeavesQty", 0)) * float(p.get("CurrentPrice", 0))
            for p in positions
        )
        order_value = shares * ref_price
        if current_value + order_value > PHASE5_MAX_CAP:
            log(f"SKIP(実注文): {code4_or_5} 残高上限超過 "
                f"(現在{current_value:,.0f}円 + 注文{order_value:,.0f}円 "
                f"> 上限{PHASE5_MAX_CAP:,.0f}円)")
            return None

        code4 = KabuClient.to_code4(code4_or_5)
        result = client.buy(code4, shares)
        order_id = result.get("OrderId", "")
        log(f"  実注文成功: {code4} x{shares}株  OrderId={order_id}")
        return order_id

    except Exception as e:
        log(f"  実注文失敗（ペーパー記録は継続）: {type(e).__name__}: {e}")
        send_error_notify(_SCRIPT, "実注文失敗", str(e))
        return None


def main() -> None:
    log(f"auto_entry.py 開始 ({'実取引' if LIVE_TRADING else 'ペーパー'}モード)")

    # 月次ストップチェック
    today_ym = date.today().strftime("%Y-%m")
    if Path(MONTHLY_STOP).exists():
        stop_ym = Path(MONTHLY_STOP).read_text(encoding="utf-8").strip()
        if stop_ym == today_ym:
            log(f"SKIP: 月次損失ストップ中 ({today_ym}) - 新規エントリーなし")
            send_notify("【ST】エントリー", f"月次損失ストップ中 - 新規エントリーなし ({today_ym})")
            return

    signals = load_today_signals()
    if signals.empty:
        log("INFO: 本日のシグナル銘柄なし")
        send_notify("【ST】エントリー", "本日のシグナルなし")
        return

    signals = signals.sort_values("RSI14", ascending=False).reset_index(drop=True)
    log(f"INFO: 本日のシグナル {len(signals)} 件")

    wb       = load_workbook(TRADE_LOG)
    ws_trade = wb["取引記録"]
    config   = read_config(wb)

    max_pos  = int(config.get("最大同時保有", MAX_POSITIONS))
    max_stk  = float(config.get("1銘柄上限", MAX_PER_STOCK))

    holdings       = get_current_holdings(ws_trade)
    available      = max_pos - len(holdings)
    held_codes     = {h["code"] for h in holdings}
    log(f"INFO: 現保有 {len(holdings)} 銘柄 / 上限 {max_pos} / 空き {available} 枠")

    if available <= 0:
        log("INFO: 保有上限に達しているためエントリーなし")
        wb.close()
        return

    today        = date.today()
    next_row     = get_next_empty_row(ws_trade)
    entered      = 0
    entries_info = []   # (code, rsi) for LINE notification

    for _, sig in signals.iterrows():
        if entered >= available:
            break

        code        = str(sig["Code"]).zfill(4)
        strategy    = str(sig.get("Strategy", "A")).strip()
        entry_price = float(sig["Close"])
        stop_loss   = float(sig["StopLoss"])
        take_profit = float(sig["TakeProfit"])
        atr  = float(sig["ATR14"]) if pd.notna(sig.get("ATR14"))  else None
        rsi  = float(sig["RSI14"]) if pd.notna(sig.get("RSI14"))  else None
        adx  = float(sig["ADX14"]) if pd.notna(sig.get("ADX14"))  else None

        if code in held_codes:
            log(f"SKIP: {code} は既に保有中")
            continue

        # 1銘柄上限: 1株でも購入できるか確認
        if entry_price > max_stk:
            log(f"SKIP: {code} 株価 {entry_price:,.0f}円 > 上限 {max_stk:,.0f}円")
            continue

        ws_trade.cell(row=next_row, column=1,  value=today)
        ws_trade.cell(row=next_row, column=2,  value=code)
        ws_trade.cell(row=next_row, column=3,  value=entry_price)
        ws_trade.cell(row=next_row, column=4,  value=stop_loss)
        ws_trade.cell(row=next_row, column=5,  value=take_profit)
        ws_trade.cell(row=next_row, column=6,  value=atr)
        ws_trade.cell(row=next_row, column=7,  value=rsi)
        ws_trade.cell(row=next_row, column=8,  value=adx)
        ws_trade.cell(row=next_row, column=9,  value=None)
        ws_trade.cell(row=next_row, column=10, value=None)
        ws_trade.cell(row=next_row, column=11, value=None)
        ws_trade.cell(row=next_row, column=12, value=None)
        ws_trade.cell(row=next_row, column=13, value="保有中")
        ws_trade.cell(row=next_row, column=14, value=strategy)

        hold_limit = {"A": 10, "C": 7, "D": 5, "E": 10, "F": 15}.get(strategy, 10)
        shares = max(1, int(max_stk // entry_price))
        log(f"ENTRY: [{strategy}] {code} @ {entry_price:,.0f}円  株数={shares}  "
            f"SL={stop_loss:,.0f}  TP={take_profit:,.0f}  "
            f"RSI={rsi:.1f}  保有上限={hold_limit}日")

        # ── Phase 5: 実注文 ──────────────────────────────────
        order_id = None
        if LIVE_TRADING:
            order_id = _place_live_buy(code, shares, entry_price)

        entries_info.append((code, rsi, strategy, order_id))
        held_codes.add(code)
        next_row += 1
        entered  += 1

    wb.save(TRADE_LOG)
    log(f"INFO: {entered} 件エントリー記録 → {TRADE_LOG}")

    # 通知（エントリー結果サマリー）
    total_held = len(holdings) + entered
    mode_str   = "【実取引】" if LIVE_TRADING else "【ペーパー】"
    body_lines = [f"{mode_str} 新規：{entered}件"]
    for code, rsi, strat, order_id in entries_info:
        rsi_str = f"{rsi:.1f}" if rsi is not None else "N/A"
        oid_str = f" OrderID:{order_id}" if order_id else ""
        body_lines.append(f"　[{strat}] {code} RSI {rsi_str}{oid_str}")
    body_lines.append(f"保有中：{total_held}銘柄")
    send_notify("【ST】エントリー記録完了", "\n".join(body_lines))


if __name__ == "__main__":
    try:
        main()
    except PermissionError as e:
        log(f"ERROR: PermissionError: {e}")
        send_error_notify(_SCRIPT, "PermissionError", str(e))
        sys.exit(1)
    except FileNotFoundError as e:
        log(f"ERROR: FileNotFoundError: {e}")
        send_error_notify(_SCRIPT, "FileNotFoundError", str(e))
        sys.exit(1)
    except ValueError as e:
        log(f"ERROR: ValueError: {e}")
        send_error_notify(_SCRIPT, "ValueError", str(e))
        sys.exit(1)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        log(f"ERROR: {type(e).__name__}: {e}\n{tb}")
        send_error_notify(_SCRIPT, type(e).__name__, str(e))
        sys.exit(1)
