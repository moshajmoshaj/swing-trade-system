"""
src/auto_exit.py
Phase 4 ペーパートレード - 決済自動判定・記録
実行順序: 1番目（scanner.py より前）
"""
import sys
import os
sys.stdout.reconfigure(encoding="utf-8")

from datetime import date, datetime, timedelta
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook
import jpholiday
import jquantsapi
import sys as _sys
_sys.path.insert(0, str(Path(__file__).parent))
from notifier import send_notify, send_error_notify, call_with_retry

load_dotenv()

_SCRIPT      = "auto_exit.py"
TRADE_LOG    = "logs/paper_trade_log.xlsx"
SCHED_LOG    = "logs/scheduler_log.txt"
MONTHLY_STOP = "logs/monthly_stop.txt"

FORCED_EXIT_DAYS = 10   # 設計書準拠（10営業日）
MAX_PER_STOCK    = 200_000


def log(msg: str) -> None:
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] [exit] {msg}"
    print(line)
    os.makedirs("logs", exist_ok=True)
    with open(SCHED_LOG, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def read_config(wb) -> dict:
    ws = wb["設定"]
    config = {}
    for row in ws.iter_rows(min_row=3, max_row=15, values_only=True):
        if row[0] and row[1] is not None:
            config[str(row[0])] = row[1]
    return config


def get_client() -> jquantsapi.ClientV2:
    token = os.getenv("JQUANTS_REFRESH_TOKEN")
    if not token:
        raise EnvironmentError(".env に JQUANTS_REFRESH_TOKEN が設定されていません")
    return jquantsapi.ClientV2(api_key=token)


def business_days_held(entry_date: date, today: date) -> int:
    """entry_date 翌日から today まで（含む）の営業日数"""
    count = 0
    d = entry_date + timedelta(days=1)
    while d <= today:
        if d.weekday() < 5 and not jpholiday.is_holiday(d):
            count += 1
        d += timedelta(days=1)
    return count


def fetch_today_ohlc(client: jquantsapi.ClientV2, codes: list[str]) -> dict[str, dict]:
    """当日の High / Low / Close を銘柄コードをキーに返す。データ未取得は空辞書。"""
    today_str = date.today().strftime("%Y%m%d")
    col_rename = {"O": "Open", "H": "High", "L": "Low", "C": "Close", "Vo": "Volume"}
    prices = {}

    for code in codes:
        try:
            df = client.get_eq_bars_daily(code=code, date_yyyymmdd=today_str)
            if df is None or df.empty:
                log(f"WARN: {code} 当日価格データ未取得（市場終値未確定の可能性）")
                continue
            df = df.rename(columns=col_rename)
            row = df.iloc[-1]
            prices[code] = {
                "High":  float(row.get("High",  0)),
                "Low":   float(row.get("Low",   0)),
                "Close": float(row.get("Close", 0)),
                "Date":  str(row.get("Date", ""))[:10],
            }
        except Exception as e:
            log(f"WARN: {code} 価格取得エラー ({e})")

    return prices


def calc_monthly_pnl(ws_trade, today_ym: str) -> float:
    """当月の確定損益合計（円）を返す"""
    total = 0.0
    for row in ws_trade.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        exit_date = row[8]   # 決済日
        pnl       = row[10]  # 損益円
        if exit_date is None or pnl is None:
            continue
        if isinstance(exit_date, (date, datetime)):
            ym = exit_date.strftime("%Y-%m")
        elif isinstance(exit_date, str):
            ym = exit_date[:7]
        else:
            continue
        if ym == today_ym:
            total += float(pnl)
    return total


def main() -> None:
    log("auto_exit.py 開始")

    if not Path(TRADE_LOG).exists():
        log(f"INFO: {TRADE_LOG} が見つかりません - スキップ")
        return

    wb       = load_workbook(TRADE_LOG)
    ws_trade = wb["取引記録"]
    config   = read_config(wb)
    capital  = float(config.get("運用資金", 1_000_000))
    monthly_limit = float(config.get("月次損失上限", 0.10))

    # 保有中ポジション収集
    holdings = []
    for row in ws_trade.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        if str(row[12].value).strip() != "保有中":
            continue

        entry_date = row[0].value
        if isinstance(entry_date, datetime):
            entry_date = entry_date.date()
        elif isinstance(entry_date, str):
            entry_date = date.fromisoformat(entry_date[:10])

        holdings.append({
            "row_idx":     row[0].row,
            "code":        str(row[1].value).strip(),
            "entry_date":  entry_date,
            "entry_price": float(row[2].value),
            "stop_loss":   float(row[3].value),
            "take_profit": float(row[4].value),
        })

    if not holdings:
        log("INFO: 保有中ポジションなし")
        wb.close()
        send_notify("【ST】決済判定", "保有中ポジションなし")
        return

    log(f"INFO: 保有中 {len(holdings)} 銘柄の決済判定を開始")

    # 当日価格取得（リトライ付き・失敗時は ConnectionError を上位に伝播）
    client = call_with_retry(get_client)
    codes  = [h["code"] for h in holdings]
    prices = fetch_today_ohlc(client, codes)

    today       = date.today()
    exited      = 0
    exits_info  = []   # (code, pnl_yen, reason) for LINE notification

    for h in holdings:
        code         = h["code"]
        entry_date   = h["entry_date"]
        entry_price  = h["entry_price"]
        stop_loss    = h["stop_loss"]
        take_profit  = h["take_profit"]
        row_idx      = h["row_idx"]

        # 保有日数
        held_days = business_days_held(entry_date, today)

        price = prices.get(code)
        if price is None:
            # 当日データ未取得でも強制終了は判定（日数ベース）
            if held_days >= FORCED_EXIT_DAYS:
                log(f"WARN: {code} 当日価格未取得のため強制終了を翌日に延期")
            continue

        high  = price["High"]
        low   = price["Low"]
        close = price["Close"]

        exit_price = None
        reason     = None

        # 優先順位: 損切り > 利確 > 強制終了
        if low <= stop_loss:
            exit_price = stop_loss
            reason     = "損切り"
        elif high >= take_profit:
            exit_price = take_profit
            reason     = "利確"
        elif held_days >= FORCED_EXIT_DAYS:
            exit_price = close
            reason     = "強制終了"

        if reason is None:
            log(f"HOLD: {code} 保有{held_days}日目  H={high:,.0f}  L={low:,.0f}  "
                f"SL={stop_loss:,.0f}  TP={take_profit:,.0f}")
            continue

        # PnL 計算（200,000円配分ベース）
        shares   = max(1, int(MAX_PER_STOCK // entry_price))
        pnl_yen  = (exit_price - entry_price) * shares
        pnl_rate = (exit_price - entry_price) / entry_price

        # Excel 更新
        ws_trade.cell(row=row_idx, column=9,  value=today)
        ws_trade.cell(row=row_idx, column=10, value=exit_price)
        ws_trade.cell(row=row_idx, column=11, value=round(pnl_yen, 0))
        ws_trade.cell(row=row_idx, column=12, value=round(pnl_rate, 6))
        ws_trade.cell(row=row_idx, column=13, value=reason)

        log(f"EXIT: {code}  理由={reason}  "
            f"決済={exit_price:,.0f}  損益={pnl_yen:+,.0f}円 ({pnl_rate:+.2%})  "
            f"保有{held_days}日")
        exits_info.append((code, pnl_yen, reason))
        exited += 1

    wb.save(TRADE_LOG)
    log(f"INFO: {exited} 件決済記録 → {TRADE_LOG}")

    # Gmail 通知（決済結果サマリー）
    remaining = len(holdings) - exited
    body_lines = [f"決済：{exited}件"]
    for code, pnl, reason in exits_info:
        body_lines.append(f"　{code} {pnl:+,.0f}円（{reason}）")
    body_lines.append(f"保有中：{remaining}銘柄")
    send_notify("【ST】決済判定完了", "\n".join(body_lines))

    # 月次損失チェック（決済記録後に再集計）
    wb2      = load_workbook(TRADE_LOG)
    ws2      = wb2["取引記録"]
    today_ym = today.strftime("%Y-%m")
    monthly_pnl = calc_monthly_pnl(ws2, today_ym)
    wb2.close()

    loss_limit = -capital * monthly_limit
    log(f"INFO: 当月損益合計 {monthly_pnl:+,.0f}円  ストップライン {loss_limit:,.0f}円")

    if monthly_pnl < loss_limit:
        Path(MONTHLY_STOP).write_text(today_ym, encoding="utf-8")
        log(f"ALERT: 月次損失上限到達 ({monthly_pnl:+,.0f}円) - "
            f"{today_ym} の新規エントリーを停止します")
    else:
        # 前月ストップが残っていれば削除
        if Path(MONTHLY_STOP).exists():
            stop_ym = Path(MONTHLY_STOP).read_text(encoding="utf-8").strip()
            if stop_ym != today_ym:
                Path(MONTHLY_STOP).unlink()
                log(f"INFO: 前月のストップフラグ ({stop_ym}) を解除")


if __name__ == "__main__":
    os.makedirs("logs", exist_ok=True)
    with open(SCHED_LOG, "a", encoding="utf-8") as f:
        f.write("\n========================================\n")
    try:
        main()
    except PermissionError as e:
        log(f"ERROR: PermissionError: {e}")
        send_error_notify(_SCRIPT, "PermissionError", str(e))
        sys.exit(1)
    except FileNotFoundError as e:
        log(f"ERROR: FileNotFoundError: {e}")
        send_error_notify(_SCRIPT, "FileNotFoundError", str(e))
        sys.exit(1)
    except ConnectionError as e:
        log(f"ERROR: ConnectionError: {e}")
        send_error_notify(_SCRIPT, "ConnectionError", str(e))
        sys.exit(1)
    except ValueError as e:
        log(f"ERROR: ValueError: {e}")
        send_error_notify(_SCRIPT, "ValueError", str(e))
        sys.exit(1)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        log(f"ERROR: {type(e).__name__}: {e}\n{tb}")
        send_error_notify(_SCRIPT, type(e).__name__, str(e))
        sys.exit(1)
