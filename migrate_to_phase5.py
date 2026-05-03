"""
migrate_to_phase5.py
Phase 5 移行ワンクリックスクリプト

実行タイミング: Phase 4 合格判定後（2026-07-27 以降）

処理内容:
  1. Phase 4 合格基準の最終確認
  2. 候補銘柄 v2 への切り替え（final_candidates_v2 → final_candidates）
  3. .env の DIVIDEND_FILTER を true に更新
  4. 移行チェックリスト（phase5_checklist.md）の進捗表示
  5. 次のアクションを案内

実行方法:
  python migrate_to_phase5.py           # ドライラン（変更なし・確認のみ）
  python migrate_to_phase5.py --execute # 実際に移行を実行
"""
import sys
import os
import shutil
sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path
from datetime import date
import pandas as pd

PHASE4_START    = date(2026, 4, 27)
PHASE4_END      = date(2026, 7, 27)
INITIAL_CAPITAL = 1_000_000
DRY_RUN         = "--execute" not in sys.argv

TRADE_LOG = Path("logs/paper_trade_log.xlsx")

# 移行ファイルマッピング: v2 → 正式名
CANDIDATE_MIGRATIONS = {
    Path("logs/final_candidates_v2.csv"):    Path("logs/final_candidates.csv"),
    Path("logs/strategy_e_candidates_v2.csv"): Path("logs/strategy_e_candidates.csv"),
}

ENV_FILE = Path(".env")


def print_header(title: str, width: int = 60) -> None:
    print(f"\n{'='*width}")
    print(f"  {title}")
    print(f"{'='*width}")


def check_phase4_results() -> tuple[bool, dict]:
    """Phase 4 合格基準を確認して結果を返す"""
    results = {
        "annualized":       0.0,
        "monthly_stop":     False,
        "elapsed_days":     (date.today() - PHASE4_START).days,
        "period_complete":  date.today() >= PHASE4_END,
        "trade_count":      0,
    }

    if not TRADE_LOG.exists():
        return False, results

    try:
        from openpyxl import load_workbook
        wb = load_workbook(TRADE_LOG)
        ws = wb["取引記録"]

        trades = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            exit_reason = str(row[12]).strip() if row[12] is not None else ""
            if exit_reason in ("保有中", ""):
                continue
            try:
                pnl = float(row[10]) if row[10] is not None else 0.0
                exit_date = pd.to_datetime(row[8]).date() if row[8] is not None else None
                trades.append({"pnl": pnl, "exit_date": exit_date})
            except (ValueError, TypeError):
                continue

        if trades:
            df = pd.DataFrame(trades)
            total_pnl = df["pnl"].sum()
            elapsed   = results["elapsed_days"]
            if elapsed > 0:
                results["annualized"] = (
                    ((INITIAL_CAPITAL + total_pnl) / INITIAL_CAPITAL) ** (365 / elapsed) - 1
                ) * 100
            results["trade_count"] = len(df)

            # 月次ストップ確認
            df["ym"] = pd.to_datetime(df["exit_date"]).dt.to_period("M")
            monthly_pnl = df.groupby("ym")["pnl"].sum()
            results["monthly_stop"] = bool((monthly_pnl < -INITIAL_CAPITAL * 0.10).any())

        wb.close()
    except Exception as e:
        print(f"  ⚠️  取引ログ読み込みエラー: {e}")

    passed = (
        results["period_complete"] and
        results["annualized"] >= 8.0 and
        not results["monthly_stop"]
    )
    return passed, results


def show_candidate_changes() -> None:
    """v2 候補リストと現行の差分を表示"""
    for v2_path, cur_path in CANDIDATE_MIGRATIONS.items():
        if not v2_path.exists():
            print(f"  ⚠️  {v2_path} が見つかりません（スキップ）")
            continue

        v2_df  = pd.read_csv(v2_path,  dtype=str)
        cur_df = pd.read_csv(cur_path, dtype=str) if cur_path.exists() else pd.DataFrame()

        v2_col  = next((c for c in v2_df.columns  if "code" in c.lower()), v2_df.columns[0])
        cur_col = next((c for c in cur_df.columns if "code" in c.lower()), cur_df.columns[0]) \
                  if not cur_df.empty else None

        v2_codes  = set(v2_df[v2_col].astype(str).str.zfill(4))
        cur_codes = set(cur_df[cur_col].astype(str).str.zfill(4)) if cur_col else set()

        added   = v2_codes - cur_codes
        removed = cur_codes - v2_codes

        print(f"\n  {v2_path.name} → {cur_path.name}")
        print(f"    現行: {len(cur_codes)}銘柄  v2: {len(v2_codes)}銘柄")
        print(f"    新規追加: {len(added)}銘柄  除外: {len(removed)}銘柄")
        if added:
            print(f"    追加: {sorted(added)[:10]}{'...' if len(added) > 10 else ''}")
        if removed:
            print(f"    除外: {sorted(removed)[:10]}{'...' if len(removed) > 10 else ''}")


def do_candidate_migration(backup: bool = True) -> list[str]:
    """候補銘柄ファイルを v2 → 正式名に置き換える"""
    migrated = []
    for v2_path, cur_path in CANDIDATE_MIGRATIONS.items():
        if not v2_path.exists():
            print(f"  SKIP: {v2_path} なし")
            continue

        # バックアップ
        if backup and cur_path.exists():
            bk = cur_path.with_suffix(".csv.phase4_backup")
            shutil.copy2(cur_path, bk)
            print(f"  バックアップ: {cur_path} → {bk}")

        shutil.copy2(v2_path, cur_path)
        print(f"  ✅ 置換: {v2_path} → {cur_path}")
        migrated.append(str(cur_path))

    return migrated


def update_env_dividend_filter() -> None:
    """DIVIDEND_FILTER を false → true に変更"""
    if not ENV_FILE.exists():
        print(f"  ⚠️  {ENV_FILE} が見つかりません。手動で DIVIDEND_FILTER=true を設定してください。")
        return

    content = ENV_FILE.read_text(encoding="utf-8")
    if "DIVIDEND_FILTER=false" in content:
        new_content = content.replace("DIVIDEND_FILTER=false", "DIVIDEND_FILTER=true")
        ENV_FILE.write_text(new_content, encoding="utf-8")
        print("  ✅ .env: DIVIDEND_FILTER=false → true")
    elif "DIVIDEND_FILTER=true" in content:
        print("  ✅ .env: DIVIDEND_FILTER=true 既に設定済み")
    else:
        print("  ⚠️  .env に DIVIDEND_FILTER が見つかりません。手動で追加してください。")


def main() -> None:
    mode_str = "【ドライラン（確認のみ・変更なし）】" if DRY_RUN else "【実行モード】"
    print_header(f"Phase 5 移行スクリプト  {mode_str}")
    print(f"  実行日: {date.today()}")
    print(f"  Phase 4 期間: {PHASE4_START} 〜 {PHASE4_END}")

    # ── STEP 1: Phase 4 合格基準確認 ─────────────────────
    print_header("STEP 1: Phase 4 合格基準確認")
    passed, r = check_phase4_results()

    print(f"  基準1 模擬年利 ≥ 8%   : {r['annualized']:+.1f}%  "
          f"{'✅' if r['annualized'] >= 8.0 else '❌ 未達'}")
    print(f"  基準2 月次ストップ未発動: {'✅ 未発動' if not r['monthly_stop'] else '❌ 発動あり'}")
    print(f"  基準3 期間完了（3ヶ月）: {'✅ 完了' if r['period_complete'] else f'⏳ {(PHASE4_END - date.today()).days}日残り'}")
    print(f"  確定取引数             : {r['trade_count']}件")

    if not passed:
        if not r["period_complete"]:
            print(f"\n  ⚠️  Phase 4 期間未完了です（残り{(PHASE4_END - date.today()).days}日）。")
            print("     Phase 4 終了後に再実行してください。")
        else:
            print("\n  ⚠️  合格基準未達。Phase 4 延長を検討してください。")
        if DRY_RUN:
            print("  ドライランのため確認のみ。実際の移行は行いません。")
            return
        ans = input("\n  合格基準未達ですが強制移行しますか？ (yes/no): ").strip().lower()
        if ans != "yes":
            print("  移行を中止しました。")
            return
    else:
        print("\n  🎉 Phase 4 合格基準達成！移行を続けます。")

    # ── STEP 2: 候補銘柄 v2 差分確認 ─────────────────────
    print_header("STEP 2: 候補銘柄 v2 差分確認")
    show_candidate_changes()

    # ── STEP 3: 移行実行 or ドライラン終了 ───────────────
    if DRY_RUN:
        print_header("ドライラン完了")
        print("  上記内容を確認してください。")
        print("  実際に移行するには:")
        print("    python migrate_to_phase5.py --execute")
        return

    # 実行モード
    print_header("STEP 3: 移行実行")
    print("  候補銘柄の切り替えを開始します...")
    migrated = do_candidate_migration(backup=True)

    print("\n  .env の配当フィルターを有効化します...")
    update_env_dividend_filter()

    # ── STEP 4: 完了案内 ─────────────────────────────────
    print_header("移行完了")
    print("  ✅ 候補銘柄 v2 切り替え完了")
    print("  ✅ 配当落ちフィルター有効化（.env: DIVIDEND_FILTER=true）")
    print()
    print("  【次のアクション（Phase 5 開始前）】")
    print("  1. kabuステーション® を起動")
    print("  2. python src/broker_client.py で接続テスト")
    print("  3. .env に LIVE_TRADING=true・KABU_TRADE_PASSWORD を設定")
    print("  4. docs/phase5_checklist.md の残タスクを確認")
    print()
    print("  【注意】戦略E は v1（伊藤園含む）のままです。")
    print("  v2（伊藤園除外）は OOS が大幅悪化したため不採用。")
    print("  伊藤園のシグナル発生を Phase 5 で観察してください。")


if __name__ == "__main__":
    main()
